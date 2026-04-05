import io
import re
import openpyxl
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils.translation import gettext as _
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from stock.models import Article, Devise, SousTypeArticle, Unite, Stock, Client, Succursale, Entreprise
from stock.serializers import EntreeSerializer, ArticleSerializer
from stock.services.article_names import normalize_nom_scientifique
import json

# Colonne optionnelle type export inventaire / listing (ex. FOAG0001) — non utilisée à la création.
_ARTICLE_CODE_RE = re.compile(r'^[A-Z]{4}\d{4}$')


def _leading_column_is_article_code(val) -> bool:
    if val is None:
        return False
    s = str(val).strip().upper().replace(' ', '')
    return bool(_ARTICLE_CODE_RE.match(s))


def _parse_optional_id_cell(cell):
    if cell is None:
        return None
    if isinstance(cell, str) and not str(cell).strip():
        return None
    try:
        return int(float(str(cell).strip()))
    except (TypeError, ValueError):
        return None


def _normalize_appro_header(val):
    if val is None:
        return ''
    s = str(val).strip().lower()
    for a, b in (('é', 'e'), ('è', 'e'), ('ê', 'e'), ('à', 'a'), ('û', 'u'), ('ô', 'o'), ('ç', 'c')):
        s = s.replace(a, b)
    return s.replace(' ', '_')


# En-têtes reconnus (normalisés) → champ logique
_APPRO_HEADER_SYNONYMS = {
    'entreprise_id': frozenset({'entreprise_id', 'reference_entreprise', 'id_entreprise'}),
    'succursale_id': frozenset({'succursale_id', 'reference_succursale', 'id_succursale'}),
    'article_id': frozenset({'article_id', 'code_article', 'code_produit'}),
    'quantite': frozenset({'quantite', 'qty', 'qte'}),
    'prix_unitaire': frozenset({'prix_unitaire', 'prix_achat'}),
    'prix_vente': frozenset({'prix_vente', 'prix_de_vente'}),
    'devise_id': frozenset({'devise_id', 'devise'}),
    'seuil_alerte': frozenset({'seuil_alerte', 'seuil'}),
    'date_expiration': frozenset({'date_expiration', 'date_exp', 'dlc'}),
}

_APPRO_LEGACY_COL = {
    'article_id': 0,
    'quantite': 1,
    'prix_unitaire': 2,
    'prix_vente': 3,
    'devise_id': 4,
    'seuil_alerte': 5,
    'date_expiration': 6,
}

_APPRO_EXT_DEFAULT_COL = {
    'entreprise_id': 0,
    'succursale_id': 1,
    'article_id': 2,
    'quantite': 3,
    'prix_unitaire': 4,
    'prix_vente': 5,
    'devise_id': 6,
    'seuil_alerte': 7,
    'date_expiration': 8,
}


def _build_appro_column_map(header_row):
    raw = [_normalize_appro_header(c) for c in (header_row or [])]
    col_map = {}
    for logical, syns in _APPRO_HEADER_SYNONYMS.items():
        for i, h in enumerate(raw):
            if h in syns:
                col_map[logical] = i
                break
    return col_map


def _appro_col_index(field, col_map, extended):
    if extended:
        if field in col_map:
            return col_map[field]
        return _APPRO_EXT_DEFAULT_COL.get(field)
    return _APPRO_LEGACY_COL.get(field)


def _appro_row_cell(row, field, col_map, extended):
    idx = _appro_col_index(field, col_map, extended)
    if idx is None or row is None or idx >= len(row):
        return None
    return row[idx]


def _parse_number_excel(cell):
    if cell is None:
        return None
    if isinstance(cell, str):
        s = cell.strip()
        if not s:
            return None
        s = s.replace(',', '.')
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(cell, (int, float)):
        return float(cell)
    return None


def _coalesce_id(first, second, third):
    for x in (first, second, third):
        if x is not None:
            return x
    return None


def _parse_devise_id_cell(devise_cell, devise_principale):
    if devise_cell is None or (isinstance(devise_cell, str) and not devise_cell.strip()):
        return devise_principale.id if devise_principale else None
    if hasattr(devise_cell, 'id'):
        return getattr(devise_cell, 'id', None)
    if isinstance(devise_cell, (int, float)):
        return int(devise_cell)
    if isinstance(devise_cell, str):
        try:
            return int(float(devise_cell.strip()))
        except ValueError:
            return None
    return None

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_template(request):
    """Télécharge un fichier Excel modèle pour l'approvisionnement."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from stock.services.tenant_context import get_tenant_ids as _get_tenant_ids_dl

    tenant_id, branch_id = _get_tenant_ids_dl(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Approvisionnement'
    entetes = [
        'entreprise_id',
        'succursale_id',
        'article_id',
        'quantite',
        'prix_unitaire',
        'prix_vente',
        'devise_id',
        'seuil_alerte',
        'date_expiration',
    ]
    ws.append(entetes)
    # Style entête
    header_fill = PatternFill(start_color='FFB300', end_color='FFB300', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="000000")
    for col in range(1, len(entetes) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    articles_qs = Article.objects.filter(entreprise_id=tenant_id) if tenant_id else Article.objects.all()
    ex_id = articles_qs.values_list('article_id', flat=True).first() or 'EXEMP0001'
    ex_dev = (
        Devise.objects.filter(entreprise_id=tenant_id, est_principal=True).values_list('id', flat=True).first()
        if tenant_id
        else Devise.objects.filter(est_principal=True).values_list('id', flat=True).first()
    )
    ex_dev = ex_dev or 1
    # Exemples : réf. entreprise / succursale = contexte JWT (modifiable si cohérent avec votre session)
    ws.append([tenant_id or '', branch_id or '', ex_id, 100, 2500, 3000, ex_dev, 10, '2026-06-30'])
    ws.append([tenant_id or '', branch_id or '', ex_id, 50, 1500, 1800, ex_dev, 5, ''])
    ws.append([tenant_id or '', branch_id or '', ex_id, 200, 500, 600, ex_dev, 20, ''])

    # Feuille de référence des articles
    ws2 = wb.create_sheet('Articles')
    ws2.append(['CODE PRODUIT', 'Nom scientifique', 'Nom commercial'])
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    for article in articles_qs:
        ws2.append([article.article_id, article.nom_scientifique, article.nom_commercial or ''])

    # Feuille de référence des devises
    ws3 = wb.create_sheet('Devises')
    ws3.append(['ID', 'Sigle', 'Nom', 'Symbole'])
    for col in range(1, 5):
        cell = ws3.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    devises_qs = Devise.objects.filter(entreprise_id=tenant_id) if tenant_id else Devise.objects.all()
    for devise in devises_qs:
        ws3.append([devise.id, devise.sigle, devise.nom, devise.symbole])

    # Feuille entreprises : IDs entreprise / succursale (toujours au moins une ligne exploitable)
    ws_ent = wb.create_sheet('entreprises')
    ws_ent.append(['entreprise_id', 'Entreprise', 'succursale_id', 'Succursale'])
    for col in range(1, 5):
        cell = ws_ent.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws_ent.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24

    def _fill_entreprises_reference_sheet():
        rows = 0
        if tenant_id:
            ent = Entreprise.objects.filter(pk=tenant_id).first()
            if ent:
                branches = list(
                    Succursale.objects.filter(entreprise_id=tenant_id, is_active=True).order_by('nom', 'id')
                )
                if branches:
                    for s in branches:
                        ws_ent.append([ent.id, ent.nom, s.id, s.nom])
                        rows += 1
                else:
                    ws_ent.append(
                        [
                            ent.id,
                            ent.nom,
                            '',
                            '— Aucune succursale en base : laisser succursale_id vide dans Approvisionnement.',
                        ]
                    )
                    rows += 1
            else:
                ws_ent.append(
                    [
                        tenant_id,
                        '— Entreprise introuvable pour cet ID (vérifiez la base).',
                        '',
                        '',
                    ]
                )
                rows += 1
            return rows
        for s in (
            Succursale.objects.select_related('entreprise')
            .filter(is_active=True)
            .order_by('entreprise_id', 'nom', 'id')
        ):
            e = s.entreprise
            ws_ent.append(
                [
                    e.id if e else '',
                    (e.nom if e else ''),
                    s.id,
                    s.nom,
                ]
            )
            rows += 1
        if rows == 0:
            for e in Entreprise.objects.all().order_by('id')[:500]:
                ws_ent.append(
                    [
                        e.id,
                        e.nom,
                        '',
                        '— Aucune succursale en base : laisser succursale_id vide.',
                    ]
                )
                rows += 1
        return rows

    _fill_entreprises_reference_sheet()

    # Feuille "Référence complète" : tous les articles (code) et devises en un seul endroit
    ws_ref = wb.create_sheet('Reference_Complete', 1)
    ref_title_fill = PatternFill(start_color='5C6BC0', end_color='5C6BC0', fill_type='solid')
    ref_title_font = Font(bold=True, color='FFFFFF', size=12)
    ref_section_fill = PatternFill(start_color='7986CB', end_color='7986CB', fill_type='solid')
    ref_section_font = Font(bold=True, color='FFFFFF')

    ws_ref.append(['RÉFÉRENCE COMPLÈTE — Utilisez UNIQUEMENT les codes/ID ci-dessous dans la feuille "Approvisionnement"'])
    ws_ref.merge_cells('A1:I1')
    ws_ref.cell(row=1, column=1).fill = ref_title_fill
    ws_ref.cell(row=1, column=1).font = ref_title_font
    ws_ref.cell(row=1, column=1).alignment = header_align
    ws_ref.row_dimensions[1].height = 22

    ws_ref.append([])
    ws_ref.append([
        'ENTREPRISE_ID — colonne 1 ; SUCCURSALE_ID — colonne 2 (voir feuille « entreprises » pour les ID et libellés). '
        'Doivent correspondre au contexte de connexion. Laisser succursale_id vide si aucune succursale.'
    ])
    ws_ref.merge_cells('A3:I3')
    ws_ref.cell(row=3, column=1).fill = ref_section_fill
    ws_ref.cell(row=3, column=1).font = ref_section_font
    ws_ref.append([])
    row_hint = ws_ref.max_row + 1
    ws_ref.append(['ARTICLE_ID (CODE) — Colonne 3 de la feuille Approvisionnement (tous les articles)'])
    ws_ref.merge_cells(f'A{row_hint}:I{row_hint}')
    ws_ref.cell(row=row_hint, column=1).fill = ref_section_fill
    ws_ref.cell(row=row_hint, column=1).font = ref_section_font
    ws_ref.append(['CODE PRODUIT (article_id)', 'Nom scientifique', 'Nom commercial'])
    hdr_art = ws_ref.max_row
    for col in range(1, 4):
        cell = ws_ref.cell(row=hdr_art, column=col)
        cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_art = hdr_art + 1
    for article in articles_qs.order_by('article_id'):
        ws_ref.append([
            article.article_id,
            article.nom_scientifique,
            article.nom_commercial or '',
        ])
        for col in range(1, 4):
            ws_ref.cell(row=row_art, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_art += 1

    ws_ref.append([])
    ws_ref.append(['DEVISE_ID — Colonne 7 de la feuille Approvisionnement (toutes les devises)'])
    r_dev_title = ws_ref.max_row
    ws_ref.merge_cells(f'A{r_dev_title}:I{r_dev_title}')
    ws_ref.cell(row=r_dev_title, column=1).fill = ref_section_fill
    ws_ref.cell(row=r_dev_title, column=1).font = ref_section_font
    ws_ref.append(['ID', 'Sigle', 'Nom', 'Symbole'])
    row_art = ws_ref.max_row
    for col in range(1, 5):
        cell = ws_ref.cell(row=row_art, column=col)
        cell.fill = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_art += 1
    for devise in devises_qs.order_by('id'):
        ws_ref.append([devise.id, devise.sigle, devise.nom, devise.symbole])
        for col in range(1, 5):
            ws_ref.cell(row=row_art, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_art += 1

    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_ref.column_dimensions[c].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modele_approvisionnement.xlsx"'
    return response

@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_approvisionnement(request):
    """Importe un fichier Excel d'approvisionnement et crée une Entree avec ses lignes."""
    from datetime import datetime
    from decimal import Decimal
    from django.db.models import Sum
    from users.models import UserBranch
    from stock.models import MouvementCaisse
    from stock.services.tenant_context import get_tenant_ids as _get_tenant_ids

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': _('Aucun fichier envoyé.')}, status=400)
    wb = openpyxl.load_workbook(file)
    ws = wb['Approvisionnement']

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    col_map = _build_appro_column_map(header_row)
    header_nonempty = [h for h in (header_row or []) if h is not None and str(h).strip()]
    extended = (
        col_map.get('entreprise_id') is not None
        or col_map.get('succursale_id') is not None
        or len(header_nonempty) >= 9
    )

    tenant_id, branch_id = _get_tenant_ids(request)
    if not tenant_id:
        return JsonResponse({'error': _('Contexte entreprise manquant. Connectez-vous et sélectionnez une entreprise.')}, status=400)

    post_ent = _parse_optional_id_cell(request.POST.get('entreprise_id'))
    post_succ = _parse_optional_id_cell(request.POST.get('succursale_id'))

    lignes = []
    targets = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue

        row = list(row)
        min_len = 9 if extended else 7
        if len(row) < min_len:
            row.extend([None] * (min_len - len(row)))

        if not extended and len(row) < 7:
            return JsonResponse(
                {
                    'error': _(
                        'Ligne %(row)d : format ancien (7 colonnes) attendu : '
                        'article_id, quantite, prix_unitaire, prix_vente, devise_id, seuil_alerte, date_expiration.'
                    )
                    % {'row': row_idx}
                },
                status=400,
            )

        row_ent = _parse_optional_id_cell(_appro_row_cell(row, 'entreprise_id', col_map, extended))
        row_succ = _parse_optional_id_cell(_appro_row_cell(row, 'succursale_id', col_map, extended))
        article_id_raw = _appro_row_cell(row, 'article_id', col_map, extended)
        quantite_raw = _appro_row_cell(row, 'quantite', col_map, extended)
        prix_unitaire_raw = _appro_row_cell(row, 'prix_unitaire', col_map, extended)
        prix_vente_raw = _appro_row_cell(row, 'prix_vente', col_map, extended)
        devise_cell = _appro_row_cell(row, 'devise_id', col_map, extended)
        seuil_alerte_raw = _appro_row_cell(row, 'seuil_alerte', col_map, extended)
        date_expiration = _appro_row_cell(row, 'date_expiration', col_map, extended)

        article_id = str(article_id_raw).strip() if article_id_raw is not None else None
        if not article_id:
            continue

        eff_ent = _coalesce_id(row_ent, post_ent, tenant_id)
        eff_succ = _coalesce_id(row_succ, post_succ, branch_id)
        targets.add((eff_ent, eff_succ))

        quantite = _parse_number_excel(quantite_raw)
        prix_unitaire = _parse_number_excel(prix_unitaire_raw)
        prix_vente = _parse_number_excel(prix_vente_raw)
        seuil_alerte = _parse_number_excel(seuil_alerte_raw)

        if prix_vente is None:
            return JsonResponse(
                {
                    'error': _(
                        'Prix de vente manquant ou invalide à la ligne %(row)d (article %(art)s). '
                        'Le prix de vente est obligatoire pour chaque ligne.'
                    )
                    % {'row': row_idx, 'art': article_id}
                },
                status=400,
            )
        if prix_vente <= 0:
            return JsonResponse(
                {
                    'error': _(
                        'Prix de vente invalide à la ligne %(row)d (article %(art)s). '
                        'Le prix de vente doit être supérieur à 0.'
                    )
                    % {'row': row_idx, 'art': article_id}
                },
                status=400,
            )

        if quantite is None:
            return JsonResponse(
                {'error': _('Quantité manquante ou invalide à la ligne %(row)d (article %(art)s).') % {'row': row_idx, 'art': article_id}},
                status=400,
            )

        # devise : résolution après validation entreprise (on re-vérifie les id plus bas)
        devise_id_val = _parse_devise_id_cell(devise_cell, None)

        date_str = None
        if date_expiration:
            if isinstance(date_expiration, str):
                try:
                    date_str = datetime.strptime(date_expiration, '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    try:
                        date_str = datetime.strptime(date_expiration, '%d/%m/%Y').strftime('%Y-%m-%d')
                    except ValueError:
                        date_str = None
            elif isinstance(date_expiration, datetime):
                date_str = date_expiration.strftime('%Y-%m-%d')
            elif isinstance(date_expiration, (int, float)):
                try:
                    from openpyxl.utils.datetime import from_excel

                    date_obj = from_excel(date_expiration)
                    date_str = date_obj.strftime('%Y-%m-%d')
                except Exception:
                    date_str = None

        ligne = {
            'article_id': article_id,
            'quantite': int(quantite) if quantite is not None else None,
            'prix_unitaire': float(prix_unitaire) if prix_unitaire is not None else 0.0,
            'prix_vente': float(prix_vente),
            'seuil_alerte': int(seuil_alerte) if seuil_alerte is not None else 0,
            'date_expiration': date_str if date_str else None,
            'devise_id': devise_id_val,
            '_row': row_idx,
            '_eff_ent': eff_ent,
            '_eff_succ': eff_succ,
        }
        lignes.append(ligne)

    if not lignes:
        return JsonResponse({'error': _('Aucune ligne d\'approvisionnement valide dans le fichier.')}, status=400)

    if len(targets) > 1:
        return JsonResponse(
            {
                'error': _(
                    'Toutes les lignes doivent cibler la même entreprise et la même succursale '
                    '(après remplissage des colonnes ou des champs entreprise_id / succursale_id du formulaire).'
                )
            },
            status=400,
        )

    final_ent, final_succ = next(iter(targets))
    if final_ent is None:
        return JsonResponse({'error': _('Entreprise cible introuvable pour cet import.')}, status=400)
    if final_ent != tenant_id:
        return JsonResponse(
            {
                'error': _(
                    'L\'entreprise indiquée (ID %(got)d) ne correspond pas au contexte de connexion (ID %(exp)d).'
                )
                % {'got': final_ent, 'exp': tenant_id}
            },
            status=400,
        )

    if final_succ is not None:
        sc = Succursale.objects.filter(pk=final_succ, entreprise_id=final_ent).first()
        if not sc:
            return JsonResponse(
                {'error': _('Succursale ID %(sid)d invalide ou non rattachée à l\'entreprise %(eid)d.') % {'sid': final_succ, 'eid': final_ent}},
                status=400,
            )

    membership = request.user.get_current_membership(request)
    if membership and request.user.is_agent(request):
        br_qs = UserBranch.objects.filter(membership=membership, is_active=True)
        if br_qs.exists() and final_succ is not None:
            if not br_qs.filter(succursale_id=final_succ).exists():
                return JsonResponse(
                    {'error': _('Vous n\'êtes pas autorisé à approvisionner cette succursale.')},
                    status=403,
                )

    devise_principale = Devise.objects.filter(entreprise_id=final_ent, est_principal=True).first()

    for ligne in lignes:
        row_idx = ligne.pop('_row')
        ligne.pop('_eff_ent', None)
        ligne.pop('_eff_succ', None)
        aid = ligne['article_id']
        if not Article.objects.filter(pk=aid, entreprise_id=final_ent).exists():
            return JsonResponse(
                {'error': _('Ligne %(row)d : article %(art)s absent ou non rattaché à l\'entreprise %(eid)d.') % {'row': row_idx, 'art': aid, 'eid': final_ent}},
                status=400,
            )

        devise_id_val = ligne.get('devise_id')
        if devise_id_val is None:
            devise_id_val = devise_principale.id if devise_principale else None
        else:
            if not Devise.objects.filter(pk=devise_id_val, entreprise_id=final_ent).exists():
                devise_id_val = devise_principale.id if devise_principale else None
        ligne['devise_id'] = devise_id_val

    # Vérification du solde de caisse par devise avant création
    erreurs_solde = []
    totaux_par_devise = {}
    for ligne in lignes:
        devise_id = ligne.get('devise_id')
        quantite = Decimal(str(ligne.get('quantite') or 0))
        prix_unitaire = Decimal(str(ligne.get('prix_unitaire') or 0))
        montant = (quantite * prix_unitaire).quantize(Decimal('0.01'))
        if not devise_id:
            devise_id = devise_principale.id if devise_principale else None
        if not devise_id:
            continue
        if devise_id not in totaux_par_devise:
            totaux_par_devise[devise_id] = Decimal('0.00')
        totaux_par_devise[devise_id] += montant

    for devise_id, total in totaux_par_devise.items():
        devise_obj = Devise.objects.filter(pk=devise_id).first()
        if not devise_obj:
            continue
        qs = MouvementCaisse.objects.filter(devise=devise_obj, entreprise_id=final_ent)
        entrees = qs.filter(type='ENTREE').aggregate(s=Sum('montant'))['s'] or Decimal('0')
        sorties = qs.filter(type='SORTIE').aggregate(s=Sum('montant'))['s'] or Decimal('0')
        solde = entrees - sorties
        if total > solde:
            erreurs_solde.append(f"Solde insuffisant en {devise_obj.nom} ({devise_obj.sigle}): Requis {total} {devise_obj.symbole}, Disponible {solde} {devise_obj.symbole}. Veuillez d'abord effectuer une entrée en caisse dans cette devise.")

    if erreurs_solde:
        return JsonResponse({
            'soldes_insuffisants': erreurs_solde,
            'message': 'Approvisionnement impossible: soldes insuffisants dans certaines devises.'
        }, status=400)

    data = {
        'libele': request.POST.get('libele', 'Import Excel'),
        'description': request.POST.get('description', ''),
        'lignes': lignes,
    }
    serializer = EntreeSerializer(data=data, context={'request': request})
    if serializer.is_valid():
        entree = serializer.save(entreprise_id=final_ent, succursale_id=final_succ)
        from stock.services.caisse import creer_mouvement_caisse
        for devise_id, total in totaux_par_devise.items():
            if total <= 0:
                continue
            devise_obj = Devise.objects.filter(pk=devise_id).first()
            if not devise_obj:
                continue
            creer_mouvement_caisse(
                montant=total,
                devise=devise_obj,
                type_mouvement='SORTIE',
                entreprise_id=final_ent,
                succursale_id=final_succ,
                entree=entree,
                content_object=None,
                utilisateur=request.user if request.user.is_authenticated else None,
                reference_piece='',
                details=None,
                motif='',
            )
        return JsonResponse(serializer.data, status=201)
    else:
        return JsonResponse(serializer.errors, status=400)


# --- Import Articles (même logique que l'import approvisionnement) ---

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_template_articles(request):
    """Télécharge un fichier Excel modèle pour l'import d'articles."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Articles'
    entetes = ['nom_scientifique', 'nom_commercial', 'sous_type_article_id', 'unite_id', 'emplacement']
    ws.append(entetes)
    header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="000000")
    for col in range(1, len(entetes) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    # Exemples (nom_commercial et emplacement optionnels)
    ws.append(['Paracétamol 500mg', 'Doliprane', 1, 1, 'Rayon A1'])
    ws.append(['Ibuprofène 400mg', 'Advil', 1, 1, ''])
    ws.append(['Vitamine C 1g', '', 2, 2, 'Rayon B2'])

    # Feuille de référence des sous-types d'articles
    ws2 = wb.create_sheet('SousTypes')
    ws2.append(['ID', 'Libellé', 'Type article ID'])
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    for st in SousTypeArticle.objects.select_related('type_article').all():
        ws2.append([st.id, st.libelle, st.type_article_id])

    # Feuille de référence des unités
    ws3 = wb.create_sheet('Unites')
    ws3.append(['ID', 'Libellé', 'Description'])
    for col in range(1, 4):
        cell = ws3.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='FFB300', end_color='FFB300', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    for u in Unite.objects.all():
        ws3.append([u.id, u.libelle, (u.description or '')[:50]])

    # Feuille "Référence complète" : tous les ID valides en un seul endroit (aucune confusion)
    ws_ref = wb.create_sheet('Reference_Complete', 1)  # en 2e position pour être visible
    ref_title_fill = PatternFill(start_color='5C6BC0', end_color='5C6BC0', fill_type='solid')
    ref_title_font = Font(bold=True, color='FFFFFF', size=12)
    ref_section_fill = PatternFill(start_color='7986CB', end_color='7986CB', fill_type='solid')
    ref_section_font = Font(bold=True, color='FFFFFF')

    ws_ref.append(['RÉFÉRENCE COMPLÈTE — Utilisez UNIQUEMENT les ID ci-dessous dans la feuille "Articles"'])
    ws_ref.merge_cells('A1:E1')
    ws_ref.cell(row=1, column=1).fill = ref_title_fill
    ws_ref.cell(row=1, column=1).font = ref_title_font
    ws_ref.cell(row=1, column=1).alignment = header_align
    ws_ref.row_dimensions[1].height = 22

    ws_ref.append([])
    ws_ref.append(['SOUS_TYPE_ARTICLE_ID — Colonne 3 de la feuille Articles (tous les ID existants)'])
    ws_ref.merge_cells('A3:E3')
    ws_ref.cell(row=3, column=1).fill = ref_section_fill
    ws_ref.cell(row=3, column=1).font = ref_section_font
    ws_ref.append(['ID', 'Libellé sous-type', 'Type article (parent)', 'Type article ID'])
    for col in range(1, 5):
        cell = ws_ref.cell(row=4, column=col)
        cell.fill = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_sous = 5
    for st in SousTypeArticle.objects.select_related('type_article').order_by('type_article_id', 'id'):
        ws_ref.append([
            st.id,
            st.libelle,
            st.type_article.libelle if st.type_article else '',
            st.type_article_id,
        ])
        for col in range(1, 5):
            ws_ref.cell(row=row_sous, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_sous += 1

    ws_ref.append([])
    ws_ref.append(['UNITE_ID — Colonne 4 de la feuille Articles (tous les ID existants)'])
    ws_ref.merge_cells(f'A{row_sous + 1}:E{row_sous + 1}')
    ws_ref.cell(row=row_sous + 1, column=1).fill = ref_section_fill
    ws_ref.cell(row=row_sous + 1, column=1).font = ref_section_font
    row_sous += 2
    ws_ref.append(['ID', 'Libellé', 'Description'])
    for col in range(1, 4):
        cell = ws_ref.cell(row=row_sous, column=col)
        cell.fill = PatternFill(start_color='FFB300', end_color='FFB300', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_sous += 1
    for u in Unite.objects.order_by('id'):
        ws_ref.append([u.id, u.libelle, (u.description or '')])
        for col in range(1, 4):
            ws_ref.cell(row=row_sous, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_sous += 1

    for c in ['A', 'B', 'C', 'D', 'E']:
        ws_ref.column_dimensions[c].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modele_articles.xlsx"'
    return response


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_articles(request):
    """Importe un fichier Excel d'articles et crée les articles + un Stock à 0 pour chacun."""
    from stock.services.tenant_context import get_tenant_ids as _get_tenant_ids
    tenant_id, branch_id = _get_tenant_ids(request)
    if not tenant_id:
        return JsonResponse(
            {'error': _('Contexte entreprise manquant. Connectez-vous et sélectionnez une entreprise.')},
            status=400
        )

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': _('Aucun fichier envoyé.')}, status=400)

    wb = openpyxl.load_workbook(file)
    if 'Articles' not in wb.sheetnames:
        return JsonResponse(
            {'error': 'Le fichier doit contenir une feuille nommée "Articles".'},
            status=400,
        )
    ws = wb['Articles']
    created = []
    errors = []
    seen_norm_in_file = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(
            c is None or (isinstance(c, str) and not c.strip()) for c in row
        ):
            continue

        row = list(row)
        off = 1 if (row and _leading_column_is_article_code(row[0])) else 0
        if len(row) < 4 + off:
            errors.append(
                f'Ligne {row_idx}: colonnes insuffisantes. Modèle officiel : '
                f'nom_scientifique | nom_commercial | sous_type_article_id | unite_id '
                f'[| emplacement]. Si la 1re colonne est un code (ex. FOAG0001), '
                f'decalez : code (ignoré) | nom | commercial | sous_type_id | unite_id.'
            )
            continue

        nom_scientifique_raw = row[0 + off]
        nom_commercial_raw = row[1 + off] if len(row) > 1 + off else None
        sous_type_id_raw = row[2 + off] if len(row) > 2 + off else None
        unite_id_raw = row[3 + off] if len(row) > 3 + off else None
        emplacement_raw = row[4 + off] if len(row) > 4 + off else None

        nom_scientifique = (
            str(nom_scientifique_raw).strip()
            if nom_scientifique_raw is not None
            else ''
        )
        if not nom_scientifique:
            errors.append(f'Ligne {row_idx}: nom_scientifique obligatoire.')
            continue

        if _leading_column_is_article_code(nom_scientifique):
            errors.append(
                f'Ligne {row_idx}: la colonne « nom » ressemble à un code article ({nom_scientifique!r}). '
                f'Utilisez le fichier « modele_articles.xlsx » : 1re colonne = nom du produit, '
                f'ou bien une 1re colonne code + 2e colonne = nom.'
            )
            continue

        norm_batch = normalize_nom_scientifique(nom_scientifique)
        if norm_batch in seen_norm_in_file:
            errors.append(
                f'Ligne {row_idx}: nom scientifique dupliqué dans ce fichier '
                f'(même nom normalisé qu’une ligne précédente).'
            )
            continue

        nom_commercial = None
        if nom_commercial_raw is not None and str(nom_commercial_raw).strip():
            nom_commercial = str(nom_commercial_raw).strip()

        def _int_or_none(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return None
            s = str(val).strip()
            if not s:
                return None
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return None

        sous_type_id = _int_or_none(sous_type_id_raw)
        unite_id = _int_or_none(unite_id_raw)

        if not sous_type_id or not SousTypeArticle.objects.filter(pk=sous_type_id).exists():
            errors.append(
                f'Ligne {row_idx}: sous_type_article_id invalide ou inconnu ({sous_type_id_raw}).'
            )
            continue
        if not unite_id or not Unite.objects.filter(pk=unite_id).exists():
            errors.append(
                f'Ligne {row_idx}: unite_id invalide ou inconnu ({unite_id_raw}).'
            )
            continue

        emplacement = '1'
        if emplacement_raw is not None and str(emplacement_raw).strip():
            emplacement = str(emplacement_raw).strip()[:200]

        data = {
            'nom_scientifique': nom_scientifique,
            'nom_commercial': nom_commercial,
            'sous_type_article_id': sous_type_id,
            'unite_id': unite_id,
        }
        serializer = ArticleSerializer(data=data, context={'request': request})
        if not serializer.is_valid():
            errors.append(f'Ligne {row_idx}: {serializer.errors}')
            continue

        try:
            article = serializer.save(entreprise_id=tenant_id, succursale_id=branch_id)
            seen_norm_in_file.add(norm_batch)
            if emplacement != '1':
                article.emplacement = emplacement
                article.save(update_fields=['emplacement'])
            # Créer le Stock pour le nouvel article (comme ArticleViewSet.perform_create)
            Stock.objects.get_or_create(
                article=article,
                defaults={'Qte': 0, 'seuilAlert': 0},
            )
            created.append({
                'article_id': article.article_id,
                'nom_scientifique': article.nom_scientifique,
                'nom_commercial': article.nom_commercial,
            })
        except IntegrityError as e:
            err = str(e)
            if 'Duplicate entry' in err and ('stock_article' in err or 'PRIMARY' in err):
                errors.append(
                    f'Ligne {row_idx}: code article déjà existant en base (clé primaire). '
                    f'Après mise à jour du serveur, les nouveaux codes évitent les collisions entre sous-types ; '
                    f'sinon supprimez le doublon en base ou utilisez un fichier sans colonne code en double.'
                )
            else:
                errors.append(f'Ligne {row_idx}: {err}')
        except Exception as e:
            errors.append(f'Ligne {row_idx}: {str(e)}')

    if errors and not created:
        return JsonResponse({'error': 'Import échoué.', 'details': errors}, status=400)

    return JsonResponse(
        {
            'message': f'{len(created)} article(s) importé(s).',
            'crees': created,
            'erreurs': errors if errors else None,
        },
        status=201,
    )


# --- Import Sortie / Vente (même structure que approvisionnement et articles) ---

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_template_sortie(request):
    """Télécharge un fichier Excel modèle pour l'import de sorties (ventes)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tenant_id = request.user.get_entreprise_id(request) if request.user.is_authenticated else None
    articles_qs = Article.objects.filter(entreprise_id=tenant_id) if tenant_id else Article.objects.all()
    devises_qs = Devise.objects.filter(entreprise_id=tenant_id) if tenant_id else Devise.objects.all()
    clients_qs = (
        Client.objects.filter(liens_entreprise__entreprise_id=tenant_id).distinct()
        if tenant_id
        else Client.objects.all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sortie'
    entetes = ['statut', 'motif', 'client_id', 'article_id', 'quantite', 'prix_unitaire', 'devise_id']
    ws.append(entetes)
    header_fill = PatternFill(start_color='E64A19', end_color='E64A19', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="000000")
    for col in range(1, len(entetes) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # Exemples : PAYEE (payée) ou EN_CREDIT (vente à crédit — client_id obligatoire)
    ws.append(['PAYEE', 'Vente du jour', '', 'MED001', 5, 1200, 1])
    ws.append(['', '', '', 'VIT001', 2, 500, 1])
    ws.append(['EN_CREDIT', 'Vente à crédit', 'CLI0001', 'MED001', 3, 1200, 1])

    # Feuille Reference_Complete : tous les articles (code) + devises + clients
    ws_ref = wb.create_sheet('Reference_Complete', 1)
    ref_title_fill = PatternFill(start_color='5C6BC0', end_color='5C6BC0', fill_type='solid')
    ref_title_font = Font(bold=True, color='FFFFFF', size=12)
    ref_section_fill = PatternFill(start_color='7986CB', end_color='7986CB', fill_type='solid')
    ref_section_font = Font(bold=True, color='FFFFFF')

    ws_ref.append(['RÉFÉRENCE COMPLÈTE — Utilisez UNIQUEMENT les codes/ID ci-dessous dans la feuille "Sortie"'])
    ws_ref.merge_cells('A1:G1')
    ws_ref.cell(row=1, column=1).fill = ref_title_fill
    ws_ref.cell(row=1, column=1).font = ref_title_font
    ws_ref.cell(row=1, column=1).alignment = header_align
    ws_ref.row_dimensions[1].height = 22

    ws_ref.append([])
    ws_ref.append(['STATUT — Colonne 1 : PAYEE (payée) ou EN_CREDIT (vente à crédit/dette). Si EN_CREDIT, client_id (col. 3) OBLIGATOIRE.'])
    ws_ref.merge_cells('A3:G3')
    ws_ref.cell(row=3, column=1).fill = ref_section_fill
    ws_ref.cell(row=3, column=1).font = ref_section_font
    ws_ref.append([])
    ws_ref.append(['ARTICLE_ID (CODE) — Colonne 4 de la feuille Sortie (tous les articles)'])
    ws_ref.merge_cells('A5:G5')
    ws_ref.cell(row=5, column=1).fill = ref_section_fill
    ws_ref.cell(row=5, column=1).font = ref_section_font
    ws_ref.append(['CODE PRODUIT (article_id)', 'Nom scientifique', 'Nom commercial'])
    for col in range(1, 4):
        cell = ws_ref.cell(row=6, column=col)
        cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_cur = 7
    for article in articles_qs.order_by('article_id'):
        ws_ref.append([article.article_id, article.nom_scientifique, article.nom_commercial or ''])
        for col in range(1, 4):
            ws_ref.cell(row=row_cur, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_cur += 1

    ws_ref.append([])
    ws_ref.append(['DEVISE_ID — Colonne 7 de la feuille Sortie (toutes les devises)'])
    ws_ref.merge_cells(f'A{row_cur + 1}:G{row_cur + 1}')
    ws_ref.cell(row=row_cur + 1, column=1).fill = ref_section_fill
    ws_ref.cell(row=row_cur + 1, column=1).font = ref_section_font
    row_cur += 2
    ws_ref.append(['ID', 'Sigle', 'Nom', 'Symbole'])
    for col in range(1, 5):
        cell = ws_ref.cell(row=row_cur, column=col)
        cell.fill = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_cur += 1
    for devise in devises_qs.order_by('id'):
        ws_ref.append([devise.id, devise.sigle, devise.nom, devise.symbole])
        for col in range(1, 5):
            ws_ref.cell(row=row_cur, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_cur += 1

    ws_ref.append([])
    ws_ref.append(['CLIENT_ID — Colonne 3 de la feuille Sortie (OBLIGATOIRE si statut=EN_CREDIT, sinon optionnel)'])
    ws_ref.merge_cells(f'A{row_cur + 1}:G{row_cur + 1}')
    ws_ref.cell(row=row_cur + 1, column=1).fill = ref_section_fill
    ws_ref.cell(row=row_cur + 1, column=1).font = ref_section_font
    row_cur += 2
    ws_ref.append(['ID', 'Nom', 'Téléphone'])
    for col in range(1, 4):
        cell = ws_ref.cell(row=row_cur, column=col)
        cell.fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = header_align
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    row_cur += 1
    for client in clients_qs.order_by('nom'):
        ws_ref.append([client.id, client.nom, (client.telephone or '')])
        for col in range(1, 4):
            ws_ref.cell(row=row_cur, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        row_cur += 1

    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_ref.column_dimensions[c].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modele_sortie_vente.xlsx"'
    return response


@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_sortie(request):
    """Importe un fichier Excel de sortie (vente) et crée une Sortie avec ses lignes (FIFO, caisse, etc.)."""
    from stock.services.tenant_context import get_tenant_ids as _get_tenant_ids
    tenant_id, branch_id = _get_tenant_ids(request)
    if not tenant_id:
        return JsonResponse({'error': _('Contexte entreprise manquant. Connectez-vous et sélectionnez une entreprise.')}, status=400)

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'error': _('Aucun fichier envoyé.')}, status=400)

    wb = openpyxl.load_workbook(file)
    if 'Sortie' not in wb.sheetnames:
        return JsonResponse(
            {'error': 'Le fichier doit contenir une feuille nommée "Sortie".'},
            status=400,
        )

    ws = wb['Sortie']
    lignes = []
    statut_global = 'PAYEE'
    motif_global = request.POST.get('motif', 'Import Excel')
    client_id_global = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None or (isinstance(c, str) and not str(c).strip()) for c in row):
            continue

        # Colonnes : statut, motif, client_id, article_id, quantite, prix_unitaire, devise_id (min 5 pour article_id + quantite)
        if len(row) < 5:
            return JsonResponse(
                {'error': f'Ligne {row_idx}: colonnes requises statut, motif, client_id, article_id, quantite (prix_unitaire et devise_id optionnels).'},
                status=400,
            )

        statut_raw = row[0] if len(row) > 0 else None
        motif_raw = row[1] if len(row) > 1 else None
        client_id_raw = row[2] if len(row) > 2 else None
        article_id_raw = row[3] if len(row) > 3 else None
        quantite_raw = row[4] if len(row) > 4 else None
        prix_unitaire_raw = row[5] if len(row) > 5 else None
        devise_id_raw = row[6] if len(row) > 6 else None

        if row_idx == 2:
            if statut_raw is not None and str(statut_raw).strip().upper() in ('EN_CREDIT', 'EN CREDIT', 'CREDIT'):
                statut_global = 'EN_CREDIT'
            elif statut_raw is not None and str(statut_raw).strip().upper() in ('PAYEE', 'PAYÉ', 'PAYEE'):
                statut_global = 'PAYEE'
            if motif_raw is not None and str(motif_raw).strip():
                motif_global = str(motif_raw).strip()
            if client_id_raw is not None and str(client_id_raw).strip():
                try:
                    cid = str(client_id_raw).strip()
                    if Client.objects.filter(id=cid).exists():
                        client_id_global = cid
                except Exception:
                    pass

        article_id = str(article_id_raw).strip() if article_id_raw is not None else None
        if not article_id:
            continue

        def _num(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                try:
                    return int(val) if isinstance(val, float) and val == int(val) else float(val)
                except (ValueError, TypeError):
                    return None
            s = str(val).strip().replace(',', '.')
            if not s:
                return None
            try:
                return int(float(s)) if float(s) == int(float(s)) else float(s)
            except (ValueError, TypeError):
                return None

        qte = _num(quantite_raw)
        if qte is None or (isinstance(qte, (int, float)) and (qte <= 0 or (isinstance(qte, float) and qte != int(qte)))):
            return JsonResponse({'error': f'Ligne {row_idx}: quantité invalide (entier > 0).'}, status=400)
        qte = int(qte)

        prix_u = _num(prix_unitaire_raw)
        if prix_u is not None and prix_u < 0:
            return JsonResponse({'error': f'Ligne {row_idx}: prix_unitaire ne peut pas être négatif.'}, status=400)
        prix_final = float(prix_u) if prix_u is not None else 0.0

        devise_id = None
        if devise_id_raw is not None:
            if isinstance(devise_id_raw, (int, float)):
                try:
                    devise_id = int(devise_id_raw)
                except (ValueError, TypeError):
                    pass
            elif isinstance(devise_id_raw, str) and devise_id_raw.strip():
                try:
                    devise_id = int(float(devise_id_raw))
                except (ValueError, TypeError):
                    pass
        if not devise_id or not Devise.objects.filter(pk=devise_id).exists():
            default_dev = Devise.objects.filter(entreprise_id=tenant_id, est_principal=True).first()
            devise_id = default_dev.id if default_dev else None
        if not devise_id:
            return JsonResponse({'error': f'Ligne {row_idx}: devise_id invalide et aucune devise principale.'}, status=400)

        lignes.append({
            'article_id': article_id,
            'quantite': qte,
            'prix_unitaire': prix_final,
            'devise_id': devise_id,
        })

    if not lignes:
        return JsonResponse({'error': 'Aucune ligne de sortie valide dans le fichier.'}, status=400)

    # Vente en crédit (dette) : le client est obligatoire
    if statut_global == 'EN_CREDIT':
        if not client_id_global:
            return JsonResponse({
                'error': 'Pour une vente en crédit (statut EN_CREDIT), le client_id est obligatoire. Renseignez la colonne client_id sur la première ligne de la feuille Sortie.'
            }, status=400)
        try:
            client_obj = Client.objects.get(id=client_id_global)
        except Client.DoesNotExist:
            return JsonResponse({
                'error': f'Client avec ID "{client_id_global}" introuvable. Vente en crédit impossible.'
            }, status=400)

    # Appeler la même logique que SortieViewSet.create (FIFO, LigneSortieLot, Stock, MouvementCaisse)
    from rest_framework.request import Request
    from stock.views import SortieViewSet
    from stock.models import Sortie as SortieModel, DetteClient
    from decimal import Decimal

    payload = {
        'motif': motif_global,
        'statut': statut_global,
        'client_id': client_id_global,
        'lignes': lignes,
    }
    old_data = getattr(request, '_full_data', None)
    request._full_data = payload
    try:
        viewset = SortieViewSet()
        viewset.request = request
        viewset.action = 'create'
        viewset.format_kwarg = None
        viewset.kwargs = {}
        response = viewset.create(request)
    except Exception as e:
        request._full_data = old_data
        return JsonResponse({'error': f'Erreur lors de la création de la sortie: {str(e)}'}, status=400)
    request._full_data = old_data

    if response.status_code != 201:
        try:
            err_detail = response.data if hasattr(response, 'data') else str(response)
            return JsonResponse({'error': 'Création sortie refusée.', 'details': err_detail}, status=400)
        except Exception:
            return JsonResponse({'error': 'Création sortie refusée.'}, status=400)

    # Si vente en crédit : créer la DetteClient (une dette par sortie)
    if statut_global == 'EN_CREDIT' and client_id_global:
        sortie_id = response.data.get('id')
        sortie = SortieModel.objects.filter(pk=sortie_id).select_related('client').prefetch_related('lignes').first()
        if sortie and not DetteClient.objects.filter(sortie=sortie).exists():
            total_dette = Decimal('0.00')
            devise_dette = None
            for ligne in sortie.lignes.all():
                total_dette += Decimal(str(ligne.quantite)) * Decimal(str(ligne.prix_unitaire))
                if ligne.devise_id and devise_dette is None:
                    devise_dette = ligne.devise
            if devise_dette is None:
                devise_dette = Devise.objects.filter(entreprise_id=sortie.entreprise_id, est_principal=True).first()
            client_obj = Client.objects.get(id=client_id_global)
            DetteClient.objects.create(
                sortie=sortie,
                client=client_obj,
                montant_total=total_dette.quantize(Decimal('0.01')),
                devise=devise_dette,
                entreprise_id=sortie.entreprise_id,
                succursale_id=sortie.succursale_id,
            )

    return JsonResponse(response.data, status=201)
